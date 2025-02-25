import time
import logging
import pandas as pd
from datetime import datetime, timedelta
import glob
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import win32com.client

# Configure logging
logging.basicConfig(filename='error.log', level=logging.ERROR)

def perform_action(driver, element, method, offset_x, offset_y, backspace_count, text):
    """
    Perform a click action at an offset from a given element,
    send backspaces and then input the provided text.
    """
    actions = ActionChains(driver)
    if method == "by_offset":
        actions.move_by_offset(offset_x, offset_y).click()
    elif method == "to_element":
        actions.move_to_element(element).move_by_offset(offset_x, offset_y).click()
    actions.send_keys(Keys.BACKSPACE * backspace_count)
    actions.send_keys(text)
    actions.perform()

def download_csv(driver, username, password):
    # Navigate to the login page
    driver.get("http://www.ankcustoms.com/login.aspx")
    
    # Wait for username and password fields to load and fill them in
    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox1"))
    )
    username_input.send_keys(username)
    
    password_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox2"))
    )
    password_input.send_keys(password)
    
    # Submit the login form
    login_button = driver.find_element(By.ID, "Button1")
    login_button.click()
    
    # ----- First action: Input current time -----
    radio_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "rdoTrans"))
    )
    radio_button.click()
    location = radio_button.location
    size = radio_button.size
    target_x = location["x"] - 150
    target_y = location["y"] + (size["height"] / 2)
    current_time = time.strftime("%H:%M:%S")
    perform_action(driver, radio_button, "by_offset", target_x, target_y, 8, current_time)
    
    # ----- Second action: Input date (day before yesterday calculated as now - 3 days) -----
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()
    offset_x = -300
    offset_y = radio_button.size["height"] / 2
    day_before_yesterday = datetime.now() - timedelta(days=3)
    date_str = day_before_yesterday.strftime("%Y-%m-%d")
    perform_action(driver, radio_button, "to_element", offset_x, offset_y, 19, date_str)
    
    # Allow the selection to take effect
    time.sleep(1)
    
    # Click the CSV download link
    csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")
    csv_download_link.click()
    
    time.sleep(3)
    
    # Retrieve the downloaded CSV file from the Downloads folder
    downloaded_file = get_latest_downloaded_file()
    return downloaded_file

def get_latest_downloaded_file():
    downloads_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
    csv_files = glob.glob(os.path.join(downloads_folder, "*.csv"))
    if not csv_files:
        logging.error("No CSV files found in Downloads folder.")
        raise FileNotFoundError("No CSV files found in Downloads folder.")
    latest_file = max(csv_files, key=os.path.getctime)
    return latest_file

def get_csv_file(username, password):
    driver = webdriver.Chrome()
    try:
        csv_file = download_csv(driver, username, password)
    finally:
        driver.quit()
    return csv_file

def merge_and_filter_csv(csv_file1, csv_file2):
    # Read CSV files with the specified encoding
    df1 = pd.read_csv(csv_file1, encoding='cp932')
    df2 = pd.read_csv(csv_file2, encoding='cp932')
    merged_df = pd.concat([df1, df2], ignore_index=True)
    
    # Save merged CSV file with current date as filename
    merged_csv_filename = datetime.now().strftime("%Y-%m-%d.csv")
    merged_df.to_csv(merged_csv_filename, index=False)
    
    # Reload and process the merged CSV
    merged_df = pd.read_csv(merged_csv_filename)
    merged_df.iloc[:, 2] = merged_df.iloc[:, 2].str.replace("-", "")
    filtered_df = merged_df[pd.notna(merged_df.iloc[:, 11]) &
                            merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False)]
    filtered_csv_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")
    filtered_df.to_csv(filtered_csv_filename, index=False)
    
    print(f"Filtered CSV file saved as {filtered_csv_filename}")
    return filtered_csv_filename

def create_excel_from_csv(filtered_csv_filename, uof_file):
    # Extract single numbers from the third column of the filtered CSV
    filtered_data = pd.read_csv(filtered_csv_filename)
    single_numbers = filtered_data.iloc[:, 2].tolist()
    
    # Read UOF data from the provided Excel file
    uof_data = pd.read_excel(uof_file)
    
    # Create a new Excel workbook and write header row
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    column_names = ["当前时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地", "数据用"]
    for i, name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=i, value=name)
        
    # Set column widths
    col_widths = {'A': 11, 'B': 4.63, 'C': 13.25, 'D': 4.63, 'E': 18,
                  'F': 28, 'G': 58.13, 'H': 19.88, 'I': 15, 'J': 15}
    for col, width in col_widths.items():
        sheet.column_dimensions[col].width = width
        
    # Get current date in the specified format
    date_format = "%Y/%m/%d"
    current_date = datetime.now().strftime(date_format)
    
    # Fill in data from the filtered CSV starting at row 2
    row_num = 2
    for number in single_numbers:
        sheet.cell(row=row_num, column=1, value=current_date)
        sheet.cell(row=row_num, column=2, value="")
        sheet.cell(row=row_num, column=3, value=number)
        for col in range(4, 11):
            sheet.cell(row=row_num, column=col, value="")
        row_num += 1
        
    # Save the workbook with a timestamped filename
    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    excel_filename = rf"Z:\UOF\转运数据\许可\JBC{timestamp}.xlsx"
    workbook.save(excel_filename)
    
    # Perform a VLOOKUP-like operation to fill additional data from uof_data
    excel_workbook = openpyxl.load_workbook(excel_filename)
    excel_sheet = excel_workbook.active
    for row_idx, cell in enumerate(excel_sheet.iter_rows(min_row=2, max_row=excel_sheet.max_row, min_col=3, max_col=3), start=2):
        single_number = cell[0].value
        order_data = uof_data.loc[uof_data["送り状番号"].str.contains(single_number, na=False),
                                    ["箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地", "数据用"]]
        if not order_data.empty:
            vlookup_values = order_data.values.tolist()[0]
            for col_offset, value in enumerate(vlookup_values, start=4):
                excel_sheet.cell(row=row_idx, column=col_offset, value=value)
    excel_workbook.save(excel_filename)
    print("Excel 文件已创建并保存。")
    return excel_filename

def format_excel_and_export_pdf(excel_filename):
    # Start Excel via COM
    ExcelApp = win32com.client.DispatchEx("Excel.Application")
    ExcelApp.Visible = False
    abs_excel_filename = os.path.abspath(excel_filename)
    workbook = ExcelApp.Workbooks.Open(abs_excel_filename)
    worksheet = workbook.Worksheets(1)
    
    # Set row height for all used rows
    used_rows = worksheet.UsedRange.Rows.Count
    for r in range(1, used_rows + 1):
        worksheet.Rows(r).RowHeight = 30
        
    # Center align columns A to J (with specific handling for column G)
    for col in range(1, 11):
        if col != 7:
            worksheet.Columns(col).HorizontalAlignment = -4108  # Center
            worksheet.Columns(col).VerticalAlignment = -4108    # Center
    worksheet.Columns("G").VerticalAlignment = -4108  # Center
    
    # Set custom top and bottom margins (converted from centimeters to points)
    margin_points = 0.9 * 28.35
    worksheet.PageSetup.TopMargin = margin_points
    worksheet.PageSetup.BottomMargin = margin_points
    
    # Delete rows where column H is not empty (iterate backwards)
    for r in range(used_rows, 1, -1):
        if worksheet.Cells(r, 8).Value is not None:
            worksheet.Rows(r).Delete()
            
    # Add borders to all cells with data
    thin_border = 2
    used_cols = 10
    used_rows = worksheet.UsedRange.Rows.Count
    for r in range(1, used_rows + 1):
        if worksheet.Cells(r, 1).Value is not None:
            for c in range(1, used_cols + 1):
                worksheet.Cells(r, c).Borders.Weight = thin_border
                
    # Set print area and page setup options
    worksheet.PageSetup.PrintArea = f'A1:J{used_rows}'
    worksheet.PageSetup.Orientation = 2  # Landscape
    worksheet.PageSetup.PaperSize = 9    # A4
    worksheet.PageSetup.Zoom = False
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = 1
    
    workbook.Save()
    
    # Create a PDF filename based on the Excel filename timestamp
    timestamp = os.path.splitext(os.path.basename(excel_filename))[0][3:]
    pdf_filename = os.path.join(os.path.dirname(excel_filename), f'JBC-{timestamp}-许可.pdf')
    workbook.ExportAsFixedFormat(0, pdf_filename)
    
    workbook.Close(SaveChanges=False)
    ExcelApp.Quit()
    return pdf_filename

def remove_duplicates_from_excel(excel_filename):
    # Remove duplicate rows (keeping the first occurrence) based on the third column
    df = pd.read_excel(excel_filename)
    df.drop_duplicates(subset=df.columns[2], keep='first', inplace=True)
    df.to_excel(excel_filename, index=False)

def main():
    # Download CSV files using different credentials
    csv_file1 = get_csv_file("JUTB", "JUTp&HJKL2SJYjjuutt")
    csv_file2 = get_csv_file("UOFB", "EWQ&6qwe42B")
    
    # Merge and filter the CSV files
    filtered_csv_filename = merge_and_filter_csv(csv_file1, csv_file2)
    
    # Specify the UOF Excel file path
    uof_file = r"Z:\UOF\转运数据\UOF出入库汇总表.xlsx"
    
    # Create the new Excel file and perform the VLOOKUP-like operation
    excel_filename = create_excel_from_csv(filtered_csv_filename, uof_file)
    
    # Format the Excel file and export it as a PDF
    pdf_filename = format_excel_and_export_pdf(excel_filename)
    print(f"PDF 文件已创建：{pdf_filename}")
    
    # Remove duplicate rows from the Excel file
    remove_duplicates_from_excel(excel_filename)
    print("done")

if __name__ == "__main__":
    main()
