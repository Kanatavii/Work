import time
import logging
import re
import pandas as pd
from datetime import datetime, timedelta
import glob
import os
import openpyxl
import win32timezone
from selenium import webdriver
import win32com.client
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

# Configure the logging
logging.basicConfig(filename='error.log', level=logging.ERROR)

# Put the code from your first script here
def download_csv(driver, username, password):
    # Navigate to the login page
    driver.get("http://www.ankcustoms.com/login.aspx")

    # Wait for the username input field to load
    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox1"))
    )
    username_input.send_keys(username)

    # Wait for the password input field to load
    password_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox2"))
    )
    password_input.send_keys(password)

    # Submit the login form
    login_button = driver.find_element(By.ID, "Button1")
    login_button.click()
                    
    # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()

    # Get the position and size of the radio button
    radio_button_location = radio_button.location
    radio_button_size = radio_button.size

    # Calculate the target coordinates for the click
    target_x = radio_button_location["x"] - 150
    target_y = radio_button_location["y"] + (radio_button_size["height"] / 2)

    # Create an action chain
    actions = ActionChains(driver)

    # Move the mouse to the target coordinates and click
    actions.move_by_offset(target_x, target_y).click().perform()

    # Send Backspace key to delete 8 characters
    actions.send_keys(Keys.BACKSPACE * 8)

    # Get the current time
    current_time = time.strftime("%H:%M:%S")

    # Send the current time to the input field
    actions.send_keys(current_time).perform()


    # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()

    # Get the position and size of the radio button
    radio_button_location = radio_button.location
    radio_button_size = radio_button.size

    # Calculate the target coordinates for the click
    offset_x = -300
    offset_y = radio_button.size["height"] / 2

    # Create an action chain
    actions = ActionChains(driver)

    # Move the mouse to the target coordinates and click
    actions.move_to_element(radio_button).move_by_offset(offset_x, offset_y).click().perform()

    # Send Backspace key to delete 8 characters
    actions.send_keys(Keys.BACKSPACE * 19)

    # Get the current time
    now = datetime.now()

    # Calculate the date for the day before yesterday
    day_before_yesterday = datetime.now() - timedelta(days=3)
    
    # Format the date as a string in the format "YYYY-M-D"
    date_str = day_before_yesterday.strftime("%Y-%m-%d")

    # Then you can send this time to the input field
    actions.send_keys(date_str).perform()

    # Wait for the selection to take effect (adjust sleep time as needed)
    time.sleep(1)

    # Find the element for CSV下载
    csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")

    # Click the CSV下载 link
    csv_download_link.click()
    
    time.sleep(3)
    
    # Get the current URL
    current_url = driver.current_url

    # Extract the filename from the URL
    filename = os.path.basename(current_url)

    # Specify the download path and filename
    download_path = "C:\\Users\Kanat\Downloads/" + filename

    # Wait for the download to complete (adjust sleep time as needed)
    time.sleep(1)

    # Get the downloaded CSV file
    downloaded_file = get_latest_downloaded_file()

    return downloaded_file


def get_latest_downloaded_file():
    downloads_folder = os.path.expanduser('~') + '/Downloads/'
    downloaded_files = glob.glob(downloads_folder + "*.csv")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file

# Create WebDriver object
driver = webdriver.Chrome()

# Download the first CSV file
csv_file1 = download_csv(driver, "JUTB", "JUTp&HJKL2SJYjjuutt")

# Close the browser
driver.quit()

# Create WebDriver object
driver = webdriver.Chrome()

# Download the second CSV file
csv_file2 = download_csv(driver, "UOFB", "EWQ&6qwe42B")

# Close the browser
driver.quit()

# Merge the CSV files into a single DataFrame
df1 = pd.read_csv(csv_file1)
df2 = pd.read_csv(csv_file2)
merged_df = pd.concat([df1, df2], ignore_index=True)

# Create a new filename with the current date
filename = datetime.now().strftime("%Y-%m-%d.csv")

# Save the merged DataFrame to a new CSV file
merged_df.to_csv(filename, index=False)

# Load the merged CSV file
merged_df = pd.read_csv(filename)

# Replace "-" in the third column with empty string
merged_df.iloc[:, 2] = merged_df.iloc[:, 2].str.replace("-", "")

# Filter rows in the twelfth column that contain "NAKAMURA" or "GB" (case-insensitive)
filtered_df = merged_df[pd.notna(merged_df.iloc[:, 11]) & merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False)]

# Create a new filename for the filtered DataFrame with the current date
filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")

# Save the filtered DataFrame to a new CSV file
filtered_df.to_csv(filtered_filename, index=False)

print(f"Filtered CSV file saved as {filtered_filename}")

# Get the single numbers from filtered.csv
filtered_data = pd.read_csv(filtered_filename)
single_numbers = filtered_data.iloc[:, 2].tolist()

# 指定UOF出入库汇总表-7月.xlsx文件路径
uof_file = r"Z:\UOF\转运数据\UOF出入库汇总表.xlsx" 

# 读取UOF出入库汇总表-7月.xlsx文件
uof_data = pd.read_excel(uof_file)

# 创建 Excel 工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 设置列名称
column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应","入库时间","取件地"]
for i, column_name in enumerate(column_names, start=1):
    sheet.cell(row=1, column=i, value=column_name)

sheet.column_dimensions['A'].width = 11
sheet.column_dimensions['B'].width = 4.63
sheet.column_dimensions['C'].width = 13.25
sheet.column_dimensions['D'].width = 4.63
sheet.column_dimensions['E'].width = 18
sheet.column_dimensions['F'].width = 28
sheet.column_dimensions['G'].width = 58.13
sheet.column_dimensions['H'].width = 19.88
sheet.column_dimensions['I'].width = 15
    
# 设置日期格式
date_format = "%Y/%m/%d"

# 从第二行开始填充数据
row = 2

# 获取当前日期
current_date = datetime.now().strftime(date_format)

# 将数据填入相应单元格
for single_number in single_numbers:
    sheet.cell(row=row, column=1, value=current_date)
    sheet.cell(row=row, column=2, value="")
    sheet.cell(row=row, column=3, value=single_number)
    sheet.cell(row=row, column=4, value="")
    sheet.cell(row=row, column=5, value="")
    sheet.cell(row=row, column=6, value="")
    sheet.cell(row=row, column=7, value="")
    sheet.cell(row=row, column=8, value="")
    sheet.cell(row=row, column=9, value="")
    row += 1

# 保存 Excel 文件
new_filename = rf"Z:\UOF\转运数据\许可\JBC{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
workbook.save(new_filename)

# 打开Excel文件
excel_workbook = openpyxl.load_workbook(new_filename)
excel_sheet = excel_workbook.active

# 使用VLOOKUP检索对应单号的数据
for row_num, row in enumerate(excel_sheet.iter_rows(min_row=2, max_row=excel_sheet.max_row, min_col=3, max_col=3), start=2):
    single_number = row[0].value
    order_data = uof_data.loc[uof_data["送り状番号"].str.contains(single_number, na=False), ["箱数", "转运公司", "转运备注", "现场用-函数对应","入库时间","取件地"]]
    if not order_data.empty:
        vlookup_data = order_data.values.tolist()[0]
        for i, value in enumerate(vlookup_data, start=4):
            excel_sheet.cell(row=row_num, column=i, value=value)

# 保存修改过的 Excel 文件
excel_workbook.save(new_filename)

print("Excel 文件已创建并保存。")

# 创建Excel对象
Excel = win32com.client.DispatchEx("Excel.Application")
Excel.Visible = False  # Excel在后台运行

# 打开Excel文件
filename = os.path.abspath(rf"Z:\UOF\转运数据\许可\JBC{datetime.now().strftime('%Y%m%d%H%M')}.xlsx")
Workbook = Excel.Workbooks.Open(filename)
df = pd.read_excel(filename)

# 选择第一个sheet
Worksheet = Workbook.Worksheets(1)


# Set row height to 30
for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
    Worksheet.Rows(row).RowHeight = 30

# Center and middle align columns A to H except for G
for col in range(1, 10):  # Columns A to H
    if col != 7:  # Skip column G (which is the 7th column)
        Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
        Worksheet.Columns(col).VerticalAlignment = -4108  # xlCenter

# Middle align column G
Worksheet.Columns("G").VerticalAlignment = -4108  # xlCenter

# Set custom margins
margin_top_bottom = 0.9  # In centimeters
margin_points = margin_top_bottom * 28.35  # Convert centimeters to points
Worksheet.PageSetup.TopMargin = margin_points
Worksheet.PageSetup.BottomMargin = margin_points

# 删除H列非空行
for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
    if Worksheet.Cells(i, 8).Value is not None:
        Worksheet.Rows(i).Delete()

# 添加全边框
thin_border = 2
for row in range(1, Worksheet.UsedRange.Rows.Count + 1):  # 从第1行开始
    if Worksheet.Cells(row, 1).Value is not None:
        for col in range(1, 10):  # 从A到H
            Worksheet.Cells(row, col).Borders.Weight = thin_border
            
# 设置打印区域
Worksheet.PageSetup.PrintArea = 'A1:I{}'.format(Worksheet.UsedRange.Rows.Count)

# 设置为横向
Worksheet.PageSetup.Orientation = 2

# 设置为A4纸
Worksheet.PageSetup.PaperSize = 9

# 设置为适应页面
Worksheet.PageSetup.Zoom = False
Worksheet.PageSetup.FitToPagesWide = 1
Worksheet.PageSetup.FitToPagesTall = 1

# 保存更改
Workbook.Save()

# 设置输出PDF文件名
date_string = datetime.now().strftime('%Y%m%d%H%M')
output_filename = f'JBC-{date_string}-许可.pdf'

# 将Excel文件另存为PDF
Workbook.ExportAsFixedFormat(0, os.path.join(os.path.dirname(filename), output_filename))

# 退出Excel
Workbook.Close(SaveChanges=False)
Excel.Quit()

# 读取 Excel 文件
excel_filename = os.path.abspath(rf"Z:\UOF\转运数据\许可\JBC{datetime.now().strftime('%Y%m%d%H%M')}.xlsx")
df = pd.read_excel(excel_filename)

# 删除重复行，但保留第一次出现的重复值
df.drop_duplicates(subset=df.columns[2], keep='first', inplace=True)

# 保存修改后的 DataFrame 到 Excel 文件
df.to_excel(excel_filename, index=False)

print("done")