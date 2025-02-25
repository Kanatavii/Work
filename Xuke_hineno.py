import time
import logging
import re
import pandas as pd
from datetime import datetime, timedelta
import glob
import win32com.client
import os
import getpass
import openpyxl
from selenium import webdriver
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import win32com.client

# Configure the logging
logging.basicConfig(filename='error.log', level=logging.ERROR)

def download_csv(driver, username, password):
    # Navigate to the login page
    driver.get("http://www.ankcustoms.com/login.aspx")

    # Wait for the username input field to load
    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox1"))
    )
    username_input.send_keys(username)

    # Wait for the password input field to load
    password_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "TextBox2"))
    )
    password_input.send_keys(password)

    # Submit the login form
    login_button = driver.find_element(By.ID, "Button1")
    login_button.click()

    # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()

    # Get the position and size of the radio button
    radio_button_location = radio_button.location
    radio_button_size = radio_button.size

    # Calculate the target coordinates for the click
    target_x = radio_button_location["x"] - 150
    target_y = radio_button_location["y"] + (radio_button_size["height"] / 2)

    # Create an action chain
    actions = ActionChains(driver)

    # Move the mouse to the target coordinates and click
    actions.move_by_offset(target_x, target_y).click().perform()

    # Send Backspace key to delete 8 characters
    actions.send_keys(Keys.BACKSPACE * 8)

    # Get the current time
    current_time = time.strftime("%H:%M:%S")

    # Send the current time to the input field
    actions.send_keys(current_time).perform()


   # Find the element for "根据转运时间查询"
    radio_button = driver.find_element(By.ID, "rdoTrans")
    radio_button.click()

    # Get the position and size of the radio button
    radio_button_location = radio_button.location
    radio_button_size = radio_button.size

    # Calculate the target coordinates for t
    offset_x = -300
    offset_y = radio_button.size["height"] / 2

    # Create an action chain
    actions = ActionChains(driver)

    # Move the mouse to the target coordinates and click
    actions.move_to_element(radio_button).move_by_offset(offset_x, offset_y).click().perform()

    # Send Backspace key to delete 8 characters
    actions.send_keys(Keys.BACKSPACE * 19)

    # Get the current time
    now = datetime.now()

    # Calculate the date for the day before yesterday
    day_before_yesterday = datetime.now() - timedelta(days=2)
    
    # Format the date as a string in the format "YYYY-M-D"
    date_str = day_before_yesterday.strftime("%Y-%m-%d")

    # Then you can send this time to the input field
    actions.send_keys(date_str).perform()


    # Wait for the selection to take effect (adjust sleep time as needed)
    time.sleep(1)

    # Find the element for CSV下载
    csv_download_link = driver.find_element(By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")

    # Click the CSV下载 link
    csv_download_link.click()
    
    time.sleep(3)
    
    # Get the current URL
    current_url = driver.current_url

    # Extract the filename from the URL
    filename = os.path.basename(current_url)

    # Specify the download path and filename
    download_path = "C:\\Users\Kanat\Downloads/" + filename

    # Wait for the download to complete (adjust sleep time as needed)
    time.sleep(1)

    # Get the downloaded CSV file
    downloaded_file = get_latest_downloaded_file()

    return downloaded_file


def get_latest_downloaded_file():
    downloads_folder =  os.path.expanduser('~') + '/Downloads/'
    downloaded_files = glob.glob(downloads_folder + "*.csv")
    if not downloaded_files:
        raise ValueError("No CSV files found in the Downloads folder.")
    latest_file = max(downloaded_files, key=os.path.getctime)
    return latest_file


# Create WebDriver object
driver = webdriver.Chrome()

# Download the first CSV file
csv_file1 = download_csv(driver, "JUTB", "JUTp&HJKL2SJYjjuutt")

# Close the browser
driver.quit()

# Create WebDriver object
driver = webdriver.Chrome()

# Download the second CSV file
csv_file2 = download_csv(driver, "UOFB", "EWQ&6qwe42B")

# Close the browser
driver.quit()

# Merge the CSV files into a single DataFrame
df1 = pd.read_csv(csv_file1)
df2 = pd.read_csv(csv_file2)
merged_df = pd.concat([df1, df2], ignore_index=True)
missing_values = merged_df.isna().sum()

# Create a new filename with the current date
filename = datetime.now().strftime("%Y-%m-%d.csv")

# Save the merged DataFrame to a new CSV file
merged_df.to_csv(filename, index=False)
merged_df = merged_df.dropna()

# Load the merged CSV file
merged_df = pd.read_csv(filename)

# Replace "-" in the third column with empty string
merged_df.iloc[:, 2] = merged_df.iloc[:, 2].str.replace("-", "")

# Filter rows in the twelfth column that contain "NAKAMURA" or "GB" (case-insensitive)
filtered_df = merged_df[merged_df.iloc[:, 11].str.contains("NAKAMURA|GB", case=False, na=False)]

# Create a new filename for the filtered DataFrame with the current date
filtered_filename = datetime.now().strftime("%Y-%m-%d_filtered.csv")

# Save the filtered DataFrame to a new CSV file
filtered_df.to_csv(filtered_filename, index=False)

print(f"Filtered CSV file saved as {filtered_filename}")

# Get the single numbers from filtered.csv
filtered_data = pd.read_csv(filtered_filename)
single_numbers = filtered_data.iloc[:, 2].tolist()

# 初始化Chrome浏览器
driver = webdriver.Chrome()

# 导航到指定的URL
driver.get("https://quickconnect.to/")

# 在页面上找到QuickConnect ID输入框并输入信息
input_box = driver.find_element(By.ID, "input-id")
input_box.send_keys("uof-jp")

# 使用显式等待确保提交按钮已经可点击（即不再是禁用状态）
wait = WebDriverWait(driver, 60)  # 等待最长20秒
submit_button = wait.until(EC.element_to_be_clickable((By.ID, "input-submit")))
submit_button.click()

# 找到用户名输入框
wait = WebDriverWait(driver, 60)  # 等待最长20秒
username_box = wait.until(EC.presence_of_element_located((By.NAME, "username")))
username_box.send_keys("anguri")  # 输入用户名
login_button = driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]")
login_button.click()

# 找到密码输入框并输入信息
wait = WebDriverWait(driver, 60)  # 等待最长20秒
password_box = wait.until(EC.presence_of_element_located((By.NAME, "current-password")))
password_box.send_keys("uofjpA-1")
login_button = driver.find_element(By.XPATH, "//div[contains(@class, 'login-btn-spinner-wrapper')]")
login_button.click()

# 使用显式等待确保元素出现
wait = WebDriverWait(driver, 60)

# 点击File Station
wait = WebDriverWait(driver, 10)
first_element = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sds-desktop-shortcut"]/div/li[1]/div[1]')))
driver.execute_script("arguments[0].click();", first_element)

# 使用显式等待确保元素出现
wait = WebDriverWait(driver, 60)

# 双击UOF
action = ActionChains(driver)
uof_folder = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'UOF')]")))
action.double_click(uof_folder).perform()

# 使用显式等待确保元素出现
wait = WebDriverWait(driver, 60)

# 双击转运数据
div_element = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(),'转运数据')]")))
action.double_click(div_element).perform()

wait = WebDriverWait(driver, 20)  # 等待20秒
file_size_elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.x-grid3-cell-inner.x-grid3-col-filesize")))

print(f"Found {len(file_size_elements)} file size elements.")

# 如果没有找到文件大小元素，打印消息并退出
if not file_size_elements:
    print("No file sizes found")
    exit()

# 希望下载的文件名
desired_filename = "UOF出入库汇总表.xlsx"

# 定位所有包含文件名信息的元素
file_elements = driver.find_elements(By.XPATH, f"//div[contains(text(), '{desired_filename}')]")

# 过滤出不包含 "$" 符号的文件
filtered_file_elements = [element for element in file_elements if "$" not in element.text]

# 选择第一个匹配的文件
desired_file_element = filtered_file_elements[0] if filtered_file_elements else None

# 3. 下载指定的文件
if desired_file_element:
    print(f"Attempting to download the desired file: {desired_filename}...")
    print(desired_file_element.text)

    # 使用双击方式触发下载
    action.double_click(desired_file_element).perform()

    # 在下载后等待提示元素出现，以文件名作为判断条件
    try:
        # 等待提示元素出现
        success_message = wait.until(
            EC.presence_of_element_located((By.XPATH, f"//div[contains(text(), 'Download Complete: {desired_filename}')]"))
        )
        print("Download Successful!")
    except TimeoutException:
        print("Download did not complete within the expected time.")
    
    
def get_user_downloads_folder():
    # 获取当前登录用户的用户名
    username = getpass.getuser()

    # 构造下载文件夹的完整路径
    download_folder_path = os.path.join("C:\\Users", username, "Downloads")

    return download_folder_path

# 调用函数获取当前用户 Downloads 文件夹路径
download_folder = get_user_downloads_folder()

# 构建文件名的匹配模式
file_pattern = "UOF出入库汇总表*.xlsx"

# 获取下载文件夹中所有匹配模式的文件
matching_files = glob.glob(os.path.join(download_folder, file_pattern))

# 按文件的更新日期进行排序
matching_files.sort(key=os.path.getmtime, reverse=True)

# 选择最新的文件
if matching_files:
    latest_uof_file = matching_files[0]
    print("Latest UOF文件:", latest_uof_file)

    # 读取最新的UOF出入库汇总表.xlsx文件
    uof_data = pd.read_excel(latest_uof_file)
else:
    print("No matching UOF文件 found.")

uof_data = pd.read_excel(latest_uof_file)

# 创建 Excel 工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 设置列名称
column_names = ["许可时间", "回数", "送り状番号", "箱数", "转运公司", "转运备注", "现场用-函数对应","入库时间","取件地","数据用"]
for i, column_name in enumerate(column_names, start=1):
    sheet.cell(row=1, column=i, value=column_name)

sheet.column_dimensions['A'].width = 11
sheet.column_dimensions['B'].width = 4.63
sheet.column_dimensions['C'].width = 13.25
sheet.column_dimensions['D'].width = 4.63
sheet.column_dimensions['E'].width = 18
sheet.column_dimensions['F'].width = 25
sheet.column_dimensions['G'].width = 58.13
sheet.column_dimensions['H'].width = 19.88
sheet.column_dimensions['I'].width = 15
sheet.column_dimensions['J'].width = 15
    
# 设置日期格式
date_format = "%Y/%m/%d"

# 从第二行开始填充数据
row = 2

# 获取当前日期
current_date = datetime.now().strftime(date_format)

# 将数据填入相应单元格
for single_number in single_numbers:
    sheet.cell(row=row, column=1, value=current_date)
    sheet.cell(row=row, column=2, value="")
    sheet.cell(row=row, column=3, value=single_number)
    sheet.cell(row=row, column=4, value="")
    sheet.cell(row=row, column=5, value="")
    sheet.cell(row=row, column=6, value="")
    sheet.cell(row=row, column=7, value="")
    sheet.cell(row=row, column=8, value="")
    sheet.cell(row=row, column=9, value="")
    sheet.cell(row=row, column=10, value="")
    row += 1

# 保存 Excel 文件
excel_filename = "JBC{}.xlsx".format(datetime.now().strftime('%Y%m%d%H%M'))
workbook.save(excel_filename)

# 打开Excel文件
excel_workbook = openpyxl.load_workbook(excel_filename)
excel_sheet = excel_workbook.active

# 使用VLOOKUP检索对应单号的数据
for row_num, row in enumerate(excel_sheet.iter_rows(min_row=2, max_row=excel_sheet.max_row, min_col=3, max_col=3), start=2):
    single_number = row[0].value
    single_number_str = str(single_number)  # Convert to string
    order_data = uof_data.loc[uof_data["送り状番号"].str.contains(single_number_str, na=False), ["箱数", "转运公司", "转运备注", "现场用-函数对应","入库时间","取件地","数据用"]]
    if not order_data.empty:
        vlookup_data = order_data.values.tolist()[0]
        for i, value in enumerate(vlookup_data, start=4):
            excel_sheet.cell(row=row_num, column=i, value=value)

# 保存修改过的 Excel 文件
excel_workbook.save(excel_filename)

print("Excel 文件已创建并保存。")

# Configure the logging
logging.basicConfig(filename='error.log', level=logging.ERROR)

try:
    # Create Excel object
    Excel = win32com.client.DispatchEx("Excel.Application")
    Excel.Visible = False  # Excel runs in the background

    # Open Excel file
    filename = os.path.abspath(f"JBC{datetime.now().strftime('%Y%m%d%H%M')}.xlsx")
    Workbook = Excel.Workbooks.Open(filename)

    # Select the first sheet
    Worksheet = Workbook.Worksheets(1)

    # Set row height to 30
    for row in range(1, Worksheet.UsedRange.Rows.Count + 1):
        Worksheet.Rows(row).RowHeight = 30

    # Center and middle align columns A to G
    for col in range(1, 11):  # Columns A to J
        if col != 7:  # Skip column G (which is the 7th column)
            Worksheet.Columns(col).HorizontalAlignment = -4108  # xlCenter
            Worksheet.Columns(col).VerticalAlignment = -4108  # xlCenter

    # Middle align column G
    Worksheet.Columns("G").VerticalAlignment = -4108  # xlCenter

    # Set custom margins
    margin_top_bottom = 0.9  # In centimeters
    margin_points = margin_top_bottom * 28.35  # Convert centimeters to points
    Worksheet.PageSetup.TopMargin = margin_points
    Worksheet.PageSetup.BottomMargin = margin_points

    # Delete non-empty rows in column H
    for i in range(Worksheet.UsedRange.Rows.Count, 1, -1):
        if Worksheet.Cells(i, 8).Value is not None:
            Worksheet.Rows(i).Delete()

    # Add border to all cells
    thin_border = 2
    for row in range(1, Worksheet.UsedRange.Rows.Count + 1):  # Start from row 1
        if Worksheet.Cells(row, 1).Value is not None:
            for col in range(1, 11):  # Columns A to I
                Worksheet.Cells(row, col).Borders.Weight = thin_border

    # Set print area
    Worksheet.PageSetup.PrintArea = 'A1:J{}'.format(Worksheet.UsedRange.Rows.Count)

    # Set orientation to landscape
    Worksheet.PageSetup.Orientation = 2  # xlLandscape

    # Set paper size to A4
    try:
        Worksheet.PageSetup.PaperSize = 9  # xlPaperA4
    except Exception as e:
        print(f"Error setting PaperSize: {e}")
        logging.exception("Error setting PaperSize")

    # Fit to page
    Worksheet.PageSetup.Zoom = False
    Worksheet.PageSetup.FitToPagesWide = 1
    Worksheet.PageSetup.FitToPagesTall = 1

    # Save changes
    Workbook.Save()

    # Set output PDF filename
    date_string = datetime.now().strftime('%Y%m%d%H%M')
    output_filename = f'JBC-{date_string}-许可.pdf'

    # Export as PDF
    Workbook.ExportAsFixedFormat(0, os.path.join(os.path.dirname(filename), output_filename))

    # Close Excel
    Workbook.Close(SaveChanges=False)
    Excel.Quit()

    print("done")

except Exception as e:
    # Log any errors to the error.log file
    logging.exception("An error occurred:")
    print("An error occurred. Check the error.log file for details.")