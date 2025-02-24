# ======================
# IMPORTS & CONFIGURATION
# ======================
import logging
import os
import time
import glob
import pandas as pd
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl.utils import get_column_letter
import win32com.client

# Configuration
LOG_FILE = 'automation.log'
DOWNLOADS_PATH = os.path.expanduser('~') + '/Downloads/'
UOF_TEMPLATE_PATH = r"Z:\UOF\转运数据\UOF出入库汇总表.xlsx"
OUTPUT_DIR = r"Z:\UOF\转运数据\许可/"

# Logging setup
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

# ======================
# CORE FUNCTIONALITY CLASSES
# ======================

class CSVDownloader:
    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 15)
        self.actions = ActionChains(self.driver)
        
    def login(self, username, password):
        """Handle website authentication"""
        try:
            self.driver.get("http://www.ankcustoms.com/login.aspx")
            self._fill_field("TextBox1", username)
            self._fill_field("TextBox2", password)
            self.driver.find_element(By.ID, "Button1").click()
            logging.info(f"Successful login for {username}")
        except Exception as e:
            logging.error(f"Login failed: {str(e)}")
            raise

    def download_csv(self):
        """Execute CSV download workflow"""
        try:
            self._handle_radio_button()
            self._enter_time_filter()
            self._enter_date_filter()
            return self._trigger_download()
        except Exception as e:
            logging.error(f"Download failed: {str(e)}")
            raise

    def _fill_field(self, element_id, text):
        """Universal method for filling form fields"""
        field = self.wait.until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        field.clear()
        field.send_keys(text)

    def _handle_radio_button(self):
        """Handle radio button interactions"""
        radio = self.wait.until(
            EC.presence_of_element_located((By.ID, "rdoTrans"))
        )
        radio.click()
        self._smart_click_offset(radio, -150)

    def _enter_time_filter(self):
        """Input current time filter"""
        self.actions.send_keys(Keys.BACKSPACE * 8)
        self.actions.send_keys(datetime.now().strftime("%H:%M:%S"))
        self.actions.perform()

    def _enter_date_filter(self):
        """Input date filter (3 days prior)"""
        self._smart_click_offset(self.driver.find_element(By.ID, "rdoTrans"), -300)
        self.actions.send_keys(Keys.BACKSPACE * 19)
        self.actions.send_keys(
            (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d")
        )
        self.actions.perform()
        time.sleep(1)

    def _smart_click_offset(self, element, x_offset):
        """Precise offset clicking"""
        self.actions.move_to_element(element).move_by_offset(
            x_offset, element.size["height"]/2
        ).click().perform()

    def _trigger_download(self):
        """Initiate and validate CSV download"""
        download_link = self.wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "a.easyui-linkbutton[data-options*='icon-save']")
            )
        )
        download_link.click()
        time.sleep(3)
        return self._get_latest_download()

    def _get_latest_download(self):
        """Retrieve most recent download"""
        list_of_files = glob.glob(os.path.join(DOWNLOADS_PATH, "*.csv"))
        return max(list_of_files, key=os.path.getctime)

    def cleanup(self):
        """Clean up resources"""
        self.driver.quit()

class DataProcessor:
    @staticmethod
    def merge_and_filter(csv_files):
        """Process and merge CSV data"""
        dfs = []
        for f in csv_files:
            try:
                df = pd.read_csv(f, encoding='cp932')
                if not df.empty:
                    dfs.append(df)
            except Exception as e:
                logging.warning(f"Skipped invalid file {f}: {str(e)}")
        
        if not dfs:
            raise ValueError("No valid CSV data to process")
            
        merged = pd.concat(dfs, ignore_index=True)
        merged.iloc[:, 2] = merged.iloc[:, 2].str.replace("-", "")
        return merged[pd.notna(merged.iloc[:, 11]) & 
                    merged.iloc[:, 11].str.contains("NAKAMURA|GB", case=False)]

class ExcelManager:
    def __init__(self, template_path):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self._configure_columns()
        self.template_data = pd.read_excel(template_path)

    def _configure_columns(self):
        """Set up column dimensions and headers"""
        columns = [
            ("许可时间", 11), ("回数", 4.63), ("送り状番号", 13.25),
            ("箱数", 4.63), ("转运公司", 18), ("转运备注", 28),
            ("现场用-函数对应", 58.13), ("入库时间", 19.88), ("取件地", 15)
        ]
        
        for idx, (header, width) in enumerate(columns, 1):
            self.ws.cell(row=1, column=idx, value=header)
            self.ws.column_dimensions[get_column_letter(idx)].width = width

    def populate_data(self, single_numbers):
        """Fill Excel with processed data"""
        current_date = datetime.now().strftime("%Y/%m/%d")
        for row_idx, number in enumerate(single_numbers, 2):
            self.ws.cell(row=row_idx, column=1, value=current_date)
            self.ws.cell(row=row_idx, column=3, value=number)
            self._apply_vlookup(row_idx, number)

    def _apply_vlookup(self, row_idx, number):
        """Excel VLOOKUP equivalent"""
        matches = self.template_data[
            self.template_data["送り状番号"].str.contains(number, na=False)
        ]
        
        if not matches.empty:
            data = matches[["箱数", "转运公司", "转运备注", "现场用-函数对应", "入库时间", "取件地"]].values[0]
            for col_idx, value in enumerate(data, start=4):
                self.ws.cell(row=row_idx, column=col_idx, value=value)

    def save_and_format(self, filename):
        """Final formatting and save"""
        self._clean_duplicates()
        self.wb.save(filename)
        self._apply_excel_formatting(filename)
        self._convert_to_pdf(filename)
        self.wb.close()

    def _clean_duplicates(self):
        """Remove duplicate entries"""
        df = pd.DataFrame(self.ws.values)
        df.drop_duplicates(subset=df.columns[2], keep='first', inplace=True)
        for row in self.ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        for r_idx, row in enumerate(df.values, 1):
            for c_idx, value in enumerate(row, 1):
                self.ws.cell(row=r_idx, column=c_idx, value=value)

    def _apply_excel_formatting(self, filename):
        """Apply final formatting rules"""
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(filename))
            ws = wb.Worksheets(1)

            # Row formatting
            for row in range(1, ws.UsedRange.Rows.Count + 1):
                ws.Rows(row).RowHeight = 30
                for col in range(1, 10):
                    if col != 7:
                        ws.Cells(row, col).HorizontalAlignment = -4108  # xlCenter
                        ws.Cells(row, col).VerticalAlignment = -4108

            # Page setup
            ws.PageSetup.Orientation = 2  # Landscape
            ws.PageSetup.PaperSize = 9    # A4
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = 1

            # Border styling
            for row in range(1, ws.UsedRange.Rows.Count + 1):
                for col in range(1, 10):
                    ws.Cells(row, col).Borders.Weight = 2

            wb.Save()
            wb.Close()
            excel.Quit()
        except Exception as e:
            logging.error(f"Excel formatting failed: {str(e)}")
            raise

    def _convert_to_pdf(self, filename):
        """Generate PDF version"""
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(filename)
            output_pdf = filename.replace(".xlsx", ".pdf")
            wb.ExportAsFixedFormat(0, output_pdf)
            wb.Close(False)
            excel.Quit()
        except Exception as e:
            logging.error(f"PDF conversion failed: {str(e)}")
            raise

# ======================
# MAIN EXECUTION FLOW
# ======================
def main():
    try:
        # Phase 1: Data Acquisition
        downloader = CSVDownloader()
        
        # First credential set
        downloader.login("JUTB", "JUTp&HJKL2SJYjjuutt")
        csv1 = downloader.download_csv()
        downloader.cleanup()

        # Second credential set
        downloader = CSVDownloader()
        downloader.login("UOFB", "EWQ&6qwe42B")
        csv2 = downloader.download_csv()
        downloader.cleanup()

        # Phase 2: Data Processing
        processed_data = DataProcessor.merge_and_filter([csv1, csv2])
        single_numbers = processed_data.iloc[:, 2].tolist()

        # Phase 3: Report Generation
        timestamp = datetime.now().strftime("%Y%m%d%H%M")
        output_file = os.path.join(OUTPUT_DIR, f"JBC{timestamp}.xlsx")
        
        report = ExcelManager(UOF_TEMPLATE_PATH)
        report.populate_data(single_numbers)
        report.save_and_format(output_file)

        logging.info("Process completed successfully")
        print("Operation completed successfully!")

    except Exception as e:
        logging.critical(f"Critical failure: {str(e)}")
        print(f"Error occurred: {str(e)}")
        raise

if __name__ == "__main__":
    main()